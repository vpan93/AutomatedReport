import pandas as pd
from openpyxl import load_workbook, Workbook

def load_excel_sheets(file_path, sheet_names):
    """
    Load specific sheets from an Excel file into DataFrames.

    Parameters:
    file_path (str): Path to the Excel (.xlsx) file.
    sheet_names (list): List of sheet names to be loaded.

    Returns:
    tuple of pandas.DataFrame: DataFrames corresponding to the sheet names.
    """
    # Read the Excel file
    xls = pd.ExcelFile(file_path)

    # Validate the existence of the requested sheets in the Excel file
    missing_sheets = set(sheet_names) - set(xls.sheet_names)
    if missing_sheets:
        raise ValueError(f"Missing sheets: {', '.join(missing_sheets)}")

    # Load specified sheets into DataFrames
    dataframes = tuple(xls.parse(sheet_name) for sheet_name in sheet_names if sheet_name in xls.sheet_names)

    return dataframes




def write_dataframe_to_excel(existing_excel_filename, list_start_row, list_start_col, list_of_tables, list_of_titles, sheet_name):
    # Check if the Excel file exists
    if os.path.exists(existing_excel_filename):
        # Load the existing workbook
        book = load_workbook(existing_excel_filename)
    else:
        # Create a new workbook
        book = Workbook()
        # Remove the default sheet created by openpyxl
        book.remove(book.active)

    # Create a new ExcelWriter object
    writer = pd.ExcelWriter(existing_excel_filename, engine='openpyxl') 
    writer.book = book

    # Check if the sheet exists, add it if not
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    # Write the DataFrame to the sheet
    for i in range(len(list_of_tables)):
        # Writing the title for each table
        sheet = writer.book[sheet_name]
        title_cell = sheet.cell(row=list_start_row[i], column=list_start_col[i])
        title_cell.value = list_of_titles[i]

        # Write the table below the title
        table_start_row = list_start_row[i] + 1  # Start one row below the title
        list_of_tables[i].to_excel(writer, index=False, sheet_name=sheet_name, startcol=list_start_col[i], startrow=table_start_row)

    # Save the changes
    writer.save()

